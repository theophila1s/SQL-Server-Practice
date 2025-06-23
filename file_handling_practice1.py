from openpyxl import load_workbook,Workbook
wb = Workbook()
ws = wb.active
for i in range(1, 1000001):
    ws.append([f"Row {i}", f"Value {i}"])
wb.save(r"C:\Users\theop\OneDrive\Desktop\py\book2.xlsx")


from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\theop\OneDrive\Desktop\py\book1.xlsx", read_only=True)  # Use read_only mode for large files
ws = wb.active
for row in ws.iter_rows(min_row=1, max_row=10, max_col=2, values_only=True):
    print(row)  # Print the first 10 rows


from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet()
# Writing large data efficiently
for i in range(1, 1000001):
    ws.append([f"Row {i}", f"Data {i}"])
wb.save(r"C:\Users\theop\OneDrive\Desktop\py\book1.xlsx")


from openpyxl import load_workbook
wb = load_workbook(r"C:\Users\theop\OneDrive\Desktop\py\book1.xlsx", read_only=True)
ws = wb.active
# Read specific range
for row in ws.iter_rows(min_row=1, max_row=10, max_col=5, values_only=True):
    print(row)
