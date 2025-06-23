from openpyxl import load_workbook,Workbook 
wb = Workbook()
sheet = wb.active
sheet.title = "Sheet1"
sheet["A1"] = "Name"
sheet["B1"] = "Age"
sheet.append(["Shruthi", 24])
sheet.append(["Alice", 26])
wb.save(r"C:\Users\theop\OneDrive\Desktop\py\sample_ex.xlsx")

wb = load_workbook(r"C:\Users\theop\OneDrive\Desktop\py\sample_ex.xlsx")
sheet = wb.active
for row in sheet.iter_rows(values_only = True):
    print (row)